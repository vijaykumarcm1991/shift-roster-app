from urllib import response
from wsgiref import headers
from app.models import roster
from app.models.admin_user import AdminUser
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from datetime import date, timedelta
import calendar
from app.database.connection import SessionLocal
from app.models.roster import Roster
from app.models.roster_entry import RosterEntry
from app.models.employee import Employee
from app.models.shift import Shift
from jose import JWTError, jwt
from datetime import datetime, timedelta
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from passlib.context import CryptContext
from sqlalchemy.exc import ProgrammingError
from sqlalchemy import text
from collections import defaultdict
from openpyxl.comments import Comment

router = APIRouter()

SECRET_KEY = "3d060a63c0cce15bd05d6d3c91b79867ee31caeaeba4b8abd3e4cb97d4eecd32"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

security = HTTPBearer()

class EmployeeCreate(BaseModel):
    name: str
    team: str

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)

    to_encode.update({"exp": expire})

    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

class LoginRequest(BaseModel):
    username: str
    password: str

@router.post("/login")
def login(data: LoginRequest, db: Session = Depends(get_db)):

    # 👉 1. Try DB login (safe)
    try:
        admin = db.query(AdminUser).filter(
            AdminUser.username == data.username,
            AdminUser.status == "active"
        ).first()

        if admin and admin.password and pwd_context.verify(data.password, admin.password):
            token = create_access_token({"sub": data.username, "role": "admin"})
            return {"access_token": token, "token_type": "bearer"}

    except ProgrammingError:
        db.rollback()  # 🔥 VERY IMPORTANT

    # 👉 2. Fallback (ALWAYS WORKS)
    if data.username == "admin" and data.password == "admin123":
        token = create_access_token({"sub": data.username, "role": "admin"})
        return {"access_token": token, "token_type": "bearer"}

    raise HTTPException(status_code=401, detail="Invalid credentials")

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):

    token = credentials.credentials

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload

    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

@router.post("/rosters")
def create_roster(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):

    # 🔒 ADMIN CHECK (ADD HERE - FIRST LINE)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    # ❌ prevent duplicate roster
    existing = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Roster already exists")

    # create roster record
    roster = Roster(month=month, year=year)
    db.add(roster)
    db.commit()
    db.refresh(roster)

    employees = db.query(Employee).filter(Employee.status == "active").all()

    days_in_month = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)

    entries = []

    for emp in employees:
        for i in range(days_in_month):
            roster_date = start_date + timedelta(days=i)

            entry = RosterEntry(
                roster_id=roster.id,
                employee_id=emp.id,
                date=roster_date,
                shift_id=None
            )

            entries.append(entry)

    db.bulk_save_objects(entries)
    db.commit()

    return {
        "message": "Roster created successfully",
        "roster_id": roster.id,
        "total_entries": len(entries)
    }

@router.get("/roster")
def get_roster(month: int, year: int, db: Session = Depends(get_db)):

    roster = db.query(Roster).filter(
        Roster.month == int(month),
        Roster.year == int(year)
    ).order_by(Roster.id.desc()).first()

    if not roster:
        return {"message": "Roster not found"}

    entries = db.query(RosterEntry).filter_by(roster_id=roster.id).all()

    result = {}

    for entry in entries:
        emp_id = entry.employee_id

        # Initialize employee
        if emp_id not in result:
            employee = db.query(Employee).filter_by(id=emp_id).first()
            result[emp_id] = {
                "employee_id": emp_id,
                "employee_name": employee.name,
                "team": employee.team,
                "shifts": {},
                "comments": {} 
            }

        # Get shift code
        shift_code = None
        if entry.shift_id:
            shift = db.query(Shift).filter_by(id=entry.shift_id).first()
            shift_code = shift.shift_code

        # IMPORTANT: add shift (DON’T overwrite)
        result[emp_id]["shifts"][str(entry.date)] = shift_code
        result[emp_id]["comments"][str(entry.date)] = entry.comment or ""
    
    response = list(result.values())

    for item in response:
        item["status"] = roster.status  # ✅ ADD

    return response

@router.put("/roster-entry")
def update_roster_entry(
    employee_id: int,
    date: str,
    shift_code: str = "",
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    from app.models.shift import Shift

    # 👉 convert date
    roster_date = datetime.strptime(date, "%Y-%m-%d").date()

    # 👉 find entry
    entry = db.query(RosterEntry).filter_by(
        employee_id=employee_id,
        date=roster_date
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Roster entry not found")

    # 👉 OLD SHIFT
    old_shift = None
    if entry.shift_id:
        old = db.query(Shift).filter_by(id=entry.shift_id).first()
        old_shift = old.shift_code if old else None

    # ====================================================
    # 🔥 CASE 1: DELETE / CLEAR SHIFT
    # ====================================================
    if not shift_code or shift_code.strip() == "":
        entry.shift_id = None

        db.execute(text("""
            INSERT INTO audit_logs (employee_id, date, old_shift, new_shift, changed_by)
            VALUES (:emp, :date, :old, :new, :user)
        """), {
            "emp": employee_id,
            "date": roster_date,
            "old": old_shift,
            "new": None,
            "user": user.get("sub")
        })

        db.commit()

        return {"message": "Shift cleared"}

    # ====================================================
    # 🔥 CASE 2: NORMAL UPDATE
    # ====================================================
    shift = db.query(Shift).filter_by(shift_code=shift_code).first()

    if not shift:
        raise HTTPException(status_code=400, detail="Invalid shift code")

    entry.shift_id = shift.id

    db.execute(text("""
        INSERT INTO audit_logs (employee_id, date, old_shift, new_shift, changed_by)
        VALUES (:emp, :date, :old, :new, :user)
    """), {
        "emp": employee_id,
        "date": roster_date,
        "old": old_shift,
        "new": shift_code,
        "user": user.get("sub")
    })

    db.commit()

    return {"message": "Shift updated successfully"}

@router.post("/roster/bulk-update")
def bulk_update_roster(
    employee_id: int,
    start_date: str,
    end_date: str,
    shift_code: str,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    from app.models.shift import Shift

    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()

    shift = db.query(Shift).filter_by(shift_code=shift_code).first()

    if not shift:
        return {"error": "Invalid shift code"}

    updated_count = 0

    current = start
    while current <= end:

        entry = db.query(RosterEntry).filter_by(
            employee_id=employee_id,
            date=current
        ).first()

        if entry:
            # old shift
            old_shift = None
            if entry.shift_id:
                old = db.query(Shift).filter_by(id=entry.shift_id).first()
                old_shift = old.shift_code if old else None

            entry.shift_id = shift.id

            # audit log
            db.execute(text("""
                INSERT INTO audit_logs (employee_id, date, old_shift, new_shift, changed_by)
                VALUES (:emp, :date, :old, :new, :user)
            """), {
                "emp": employee_id,
                "date": current,
                "old": old_shift,
                "new": shift_code,
                "user": user.get("sub")
            })

            updated_count += 1

        current += timedelta(days=1)

    db.commit()

    return {
        "message": "Bulk update successful",
        "updated_records": updated_count
    }

@router.post("/roster/copy-week")
def copy_week(
    employee_id: int,
    source_start_date: str,
    target_start_date: str,
    db: Session = Depends(get_db)
):

    source_start = datetime.strptime(source_start_date, "%Y-%m-%d").date()
    target_start = datetime.strptime(target_start_date, "%Y-%m-%d").date()

    for i in range(7):
        source_date = source_start + timedelta(days=i)
        target_date = target_start + timedelta(days=i)

        source_entry = db.query(RosterEntry).filter_by(
            employee_id=employee_id,
            date=source_date
        ).first()

        target_entry = db.query(RosterEntry).filter_by(
            employee_id=employee_id,
            date=target_date
        ).first()

        if source_entry and target_entry:
            target_entry.shift_id = source_entry.shift_id

    db.commit()

    return {"message": "Week copied successfully"}

@router.get("/roster/summary")
def roster_summary(date: str, db: Session = Depends(get_db)):
    from app.models.shift import Shift

    target_date = datetime.strptime(date, "%Y-%m-%d").date()

    entries = db.query(RosterEntry).filter_by(date=target_date).all()

    summary = {}

    for entry in entries:
        if entry.shift_id:
            shift = db.query(Shift).filter_by(id=entry.shift_id).first()
            shift_code = shift.shift_code
        else:
            shift_code = "UNASSIGNED"

        if shift_code not in summary:
            summary[shift_code] = 0

        summary[shift_code] += 1

    return {
        "date": date,
        "summary": sorted(summary.items())
    }

@router.post("/employees")
def add_employee(
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    name = data.get("name")
    team = data.get("team")
    employee_code = data.get("employee_code")
    email = data.get("email")

    if not name or not team or not employee_code or not email:
        raise HTTPException(status_code=400, detail="Name, team, employee code, and email required")
    
    existing = db.query(Employee).filter(Employee.employee_code == employee_code).first()

    if existing:
        raise HTTPException(status_code=400, detail="Employee code already exists")

    emp = Employee(
        name=name,
        team=team,
        employee_code=employee_code,
        email=email,
        status="active"
    )

    db.add(emp)
    db.commit()
    db.refresh(emp)

    # 👉 ADD THIS BLOCK

    # get latest roster
    latest_roster = db.query(Roster).order_by(Roster.id.desc()).first()

    if latest_roster:
        import calendar
        from datetime import date, timedelta

        days_in_month = calendar.monthrange(latest_roster.year, latest_roster.month)[1]
        start_date = date(latest_roster.year, latest_roster.month, 1)

        entries = []

        for i in range(days_in_month):
            roster_date = start_date + timedelta(days=i)

            entry = RosterEntry(
                roster_id=latest_roster.id,
                employee_id=emp.id,
                date=roster_date,
                shift_id=None
            )
            entries.append(entry)

        db.bulk_save_objects(entries)
        db.commit()

    return {"message": "Employee added", "id": emp.id}

@router.put("/employees/{emp_id}")
def update_employee(
    emp_id: int,
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):

    # 🔒 ADMIN ONLY
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    emp = db.query(Employee).filter(Employee.id == emp_id).first()

    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    emp.name = data.get("name", emp.name)
    emp.team = data.get("team", emp.team)
    emp.employee_code = data.get("employee_code", emp.employee_code)
    emp.email = data.get("email", emp.email)

    db.commit()

    return {"message": "Employee updated"}

@router.delete("/employees/{emp_id}")
def delete_employee(
    emp_id: int,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    emp = db.query(Employee).filter(Employee.id == emp_id).first()

    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    emp.status = "inactive"  # ✅ soft delete
    db.commit()

    return {"message": "Employee removed"}

@router.get("/employees")
def get_employees(db: Session = Depends(get_db)):

    employees = db.query(Employee).filter(Employee.status == "active").all()

    return [
        {
            "id": emp.id,
            "name": emp.name,
            "team": emp.team,
            "employee_code": emp.employee_code,
            "email": emp.email
        }
        for emp in employees
    ]

@router.get("/roster/export")
def export_roster(month: int, year: int, db: Session = Depends(get_db)):

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        raise HTTPException(status_code=404, detail="Roster not found")

    entries = db.query(RosterEntry).filter_by(roster_id=roster.id).all()

    result = {}

    for entry in entries:
        emp_id = entry.employee_id

        if emp_id not in result:
            emp = db.query(Employee).filter_by(id=emp_id).first()
            result[emp_id] = {
                "name": emp.name,
                "team": emp.team,
                "shifts": {},
                "comments": {}
            }

        shift_code = None
        if entry.shift_id:
            shift = db.query(Shift).filter_by(id=entry.shift_id).first()
            shift_code = shift.shift_code

        result[emp_id]["shifts"][str(entry.date)] = shift_code
        result[emp_id]["comments"][str(entry.date)] = entry.comment or ""

    employees = list(result.values())

    if not employees:
        raise HTTPException(status_code=404, detail="No data found")

    dates = sorted(next(iter(result.values()))["shifts"].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # ================= STYLES =================
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    colors = {
        "S1": "ADD8E6", "S2": "FFDAB9", "S3": "DDA0DD",
        "G": "90EE90", "WO": "D3D3D3",
        "LV": "FFB6C1", "GH": "FFFFE0", "CO": "E0FFFF"
    }

    # ================= HEADER =================
    headers = ["Employee"]

    for d in dates:
        dt = datetime.strptime(d, "%Y-%m-%d")
        headers.append(f"{dt.strftime('%a')}\n{dt.day}")

    headers += [""]
    headers += ["S1","S2","S3","G","WO","CO","GH","LV","WD"]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    current_row = 2

    # ================= GROUP =================
    grouped = defaultdict(list)
    for emp in employees:
        grouped[emp["team"]].append(emp)

    for team, emps in grouped.items():

        ws.cell(row=current_row, column=1, value=team).font = Font(bold=True)
        current_row += 1

        # ================= EMPLOYEE ROWS =================
        for emp in emps:

            counts = {k:0 for k in ["S1","S2","S3","G","WO","CO","GH","LV"]}
            row = [emp["name"]]

            for d in dates:
                shift = emp["shifts"][d] or "-"
                row.append(shift)

                if shift in counts:
                    counts[shift] += 1

            working_days = counts["S1"] + counts["S2"] + counts["S3"] + counts["G"]

            row += [""]
            row += [
                counts["S1"], counts["S2"], counts["S3"], counts["G"],
                counts["WO"], counts["CO"], counts["GH"], counts["LV"],
                working_days
            ]

            ws.append(row)

            for col_idx, cell in enumerate(ws[current_row], start=1):

                cell.alignment = center
                cell.border = border

                if 1 < col_idx <= len(dates)+1:
                    val = cell.value
                    current_date = dates[col_idx-2]

                    comment = emp["comments"].get(current_date)
                    if comment:
                        cell.comment = Comment(comment, "System")

                    if val in colors:
                        cell.fill = PatternFill(start_color=colors[val], end_color=colors[val], fill_type="solid")

                summary_start = len(dates) + 3
                if col_idx >= summary_start:
                    if isinstance(cell.value, int):
                        if cell.value > 2:
                            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif cell.value < 2:
                            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)

            current_row += 1

        # ================= SPACER =================
        ws.append([""])
        current_row += 1

        # ================= SHIFT SUMMARY =================
        ws.append(["Shift Summary"])
        current_row += 1

        pivot = {
            "S1":{}, "S2":{}, "S3":{}, "G":{},
            "WO":{}, "CO":{}, "GH":{}, "LV":{}
        }

        for d in dates:
            for shift in pivot:
                pivot[shift][d] = 0

        for emp in emps:
            for d in dates:
                shift = emp["shifts"][d]
                if shift in pivot:
                    pivot[shift][d] += 1

        for shift in ["S1","S2","S3","G","WO","CO","GH","LV"]:

            row = [shift]

            for d in dates:
                row.append(pivot[shift][d])

            row += [""]
            row += [""] * 9

            ws.append(row)

            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.alignment = center
                cell.border = border

                if isinstance(cell.value, int):
                    if cell.value > 2:
                        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif cell.value < 2:
                        cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

            current_row += 1

        # ================= GRAND TOTAL (FIXED) =================
        ws.append(
            ["Grand Total"] +
            [
                sum(
                    1 for emp in emps
                    if emp["shifts"][d] in ["S1","S2","S3","G"]
                )
                for d in dates
            ] +
            [""] +
            [""] * 9   # ✅ FIX: no row summary totals
        )

        current_row += 2

    # ================= AUTO WIDTH =================
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    ws.freeze_panes = "B2"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=roster_{month}_{year}.xlsx"}
    )

@router.post("/admin-users")
def add_admin(
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    username = data.get("username")
    password = data.get("password")

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username & password required")

    password = password[:72]  # 🔥 FIX bcrypt limit
    if len(password) > 72:
        raise HTTPException(status_code=400, detail="Password too long (max 72 chars)")

    hashed = pwd_context.hash(password)

    admin = AdminUser(
        username=username,
        password=hashed
    )

    db.add(admin)
    db.commit()

    return {"message": "Admin created"}

@router.get("/admin-users")
def list_admins(db: Session = Depends(get_db), user=Depends(verify_token)):

    admins = db.query(AdminUser).filter(AdminUser.status == "active").all()

    return [{"id": a.id, "username": a.username} for a in admins]

@router.delete("/admin-users/{admin_id}")
def delete_admin(
    admin_id: int,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):
    admin = db.query(AdminUser).filter(AdminUser.id == admin_id).first()

    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")

    admin.status = "inactive"
    db.commit()

    return {"message": "Admin removed"}

@router.put("/roster-entry/comment")
def add_comment(
    employee_id: int,
    date: str,
    comment: str,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):

    roster_date = datetime.strptime(date, "%Y-%m-%d").date()

    entry = db.query(RosterEntry).filter_by(
        employee_id=employee_id,
        date=roster_date
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    entry.comment = comment

    db.execute(text("""
        INSERT INTO audit_logs (employee_id, date, old_shift, new_shift, changed_by)
        VALUES (:emp, :date, :old, :new, :user)
    """), {
        "emp": employee_id,
        "date": roster_date,
        "old": "COMMENT",
        "new": comment,
        "user": user.get("sub")
    })

    db.commit()

    return {"message": "Comment added"}

@router.get("/audit-logs")
def get_audit_logs(month: int, year: int, db: Session = Depends(get_db)):

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        return []

    logs = db.execute(text("""
        SELECT 
            a.employee_id,
            e.name,
            a.date,
            a.old_shift,
            a.new_shift,
            a.changed_by,
            a.changed_at
        FROM audit_logs a
        JOIN employees e ON e.id = a.employee_id
        ORDER BY a.changed_at DESC
        LIMIT 200
    """)).fetchall()

    result = []

    for row in logs:
        result.append({
            "employee": row.name,
            "date": str(row.date),
            "old": row.old_shift,
            "new": row.new_shift,
            "user": row.changed_by,
            "time": row.changed_at.strftime("%Y-%m-%d %H:%M")
        })

    return result

@router.post("/roster/import")
def import_roster(
    month: int,
    year: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):

    # 🔒 ADMIN ONLY
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        raise HTTPException(status_code=400, detail="Roster not found. Create first.")

    wb = load_workbook(file.file)
    ws = wb.active

    errors = []
    updated = 0

    # 👉 Read header (dates)
    headers = [cell.value for cell in ws[1]]

    dates = []

    for h in headers[1:]:
        try:
            val = str(h).strip()

            # ✅ SUPPORT: 01-03-2026
            try:
                dt = datetime.strptime(val, "%d-%m-%Y").date()
            except:
                # ✅ SUPPORT: Excel native date format
                if isinstance(h, datetime):
                    dt = h.date()
                else:
                    break

            # 🔒 VALIDATE MONTH/YEAR
            if dt.month != int(month) or dt.year != int(year):
                raise HTTPException(
                    status_code=400,
                    detail=f"Date {val} does not match selected month/year"
                )

            dates.append(dt)

        except Exception:
            break

    employees = db.query(Employee).filter(Employee.status == "active").all()
    emp_map = {e.name: e.id for e in employees}

    valid_shifts = {"S1","S2","S3","G","WO","CO","GH","LV"}

    for row in ws.iter_rows(min_row=2, values_only=True):

        name = row[0]

        if not name or name not in emp_map:
            errors.append(f"Invalid employee: {name}")
            continue

        emp_id = emp_map[name]

        for i, entry_date in enumerate(dates):

            shift = row[i+1]

            if not shift or shift == "-":
                continue

            shift = str(shift).strip()

            if shift not in valid_shifts:
                errors.append(f"{name} {entry_date}: Invalid shift {shift}")
                continue

            entry = db.query(RosterEntry).filter_by(
                employee_id=emp_id,
                date=entry_date
            ).first()

            if not entry:
                continue

            shift_obj = db.query(Shift).filter_by(shift_code=shift).first()

            if not shift_obj:
                errors.append(f"{name} Day {day}: Shift not found in DB")
                continue

            entry.shift_id = shift_obj.id
            updated += 1

    if errors:
        raise HTTPException(status_code=400, detail="\n".join(errors[:10]))

    db.commit()

    return {
        "message": "Import successful",
        "updated": updated
    }

@router.get("/shift-allowance")
def shift_allowance(month: int, year: int, db: Session = Depends(get_db)):

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        raise HTTPException(status_code=404, detail="Roster not found")

    entries = db.query(RosterEntry).filter_by(roster_id=roster.id).all()

    result = {}

    rates = {"S1":400, "S2":400, "S3":500}

    for entry in entries:

        emp = db.query(Employee).filter_by(id=entry.employee_id).first()

        if entry.employee_id not in result:
            result[entry.employee_id] = {
                "name": emp.name,
                "code": emp.employee_code,
                "email": emp.email,
                "S1":0,
                "S2":0,
                "S3":0
            }

        if entry.shift_id:
            shift = db.query(Shift).filter_by(id=entry.shift_id).first()
            if shift.shift_code in ["S1","S2","S3"]:
                result[entry.employee_id][shift.shift_code] += 1

    # 🔥 CALCULATE ALLOWANCE
    final = []

    totals = {
        "S1":0,"S2":0,"S3":0,
        "S1_amt":0,"S2_amt":0,"S3_amt":0,
        "grand":0
    }

    for emp in result.values():

        s1_amt = emp["S1"] * rates["S1"]
        s2_amt = emp["S2"] * rates["S2"]
        s3_amt = emp["S3"] * rates["S3"]

        grand = s1_amt + s2_amt + s3_amt

        totals["S1"] += emp["S1"]
        totals["S2"] += emp["S2"]
        totals["S3"] += emp["S3"]

        totals["S1_amt"] += s1_amt
        totals["S2_amt"] += s2_amt
        totals["S3_amt"] += s3_amt
        totals["grand"] += grand

        final.append({
            "name": emp["name"],
            "code": emp["code"],
            "email": emp["email"],
            "S1": emp["S1"],
            "S2": emp["S2"],
            "S3": emp["S3"],
            "S1_amt": s1_amt,
            "S2_amt": s2_amt,
            "S3_amt": s3_amt,
            "grand": grand
        })

    return {
        "employees": final,
        "totals": totals
    }

@router.get("/shift-allowance/export")
def export_shift_allowance(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    user=Depends(verify_token)
):

    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        raise HTTPException(status_code=404, detail="Roster not found")

    entries = db.query(RosterEntry).filter_by(roster_id=roster.id).all()

    result = {}

    rates = {"S1":400, "S2":400, "S3":500}

    for entry in entries:

        emp = db.query(Employee).filter_by(id=entry.employee_id).first()

        if entry.employee_id not in result:
            result[entry.employee_id] = {
                "name": emp.name,
                "code": emp.employee_code,
                "email": emp.email,
                "S1":0,"S2":0,"S3":0
            }

        if entry.shift_id:
            shift = db.query(Shift).filter_by(id=entry.shift_id).first()
            if shift.shift_code in ["S1","S2","S3"]:
                result[entry.employee_id][shift.shift_code] += 1

    # ================= EXCEL =================
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Allowance"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        "Name","Employee Code","Email",
        "S1","S2","S3",
        "S1 ₹","S2 ₹","S3 ₹",
        "Grand Total"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    totals = {
        "S1":0,"S2":0,"S3":0,
        "S1_amt":0,"S2_amt":0,"S3_amt":0,
        "grand":0
    }

    row_idx = 2

    for emp in result.values():

        s1 = emp["S1"]
        s2 = emp["S2"]
        s3 = emp["S3"]

        s1_amt = s1 * rates["S1"]
        s2_amt = s2 * rates["S2"]
        s3_amt = s3 * rates["S3"]

        grand = s1_amt + s2_amt + s3_amt

        totals["S1"] += s1
        totals["S2"] += s2
        totals["S3"] += s3
        totals["S1_amt"] += s1_amt
        totals["S2_amt"] += s2_amt
        totals["S3_amt"] += s3_amt
        totals["grand"] += grand

        ws.append([
            emp["name"], emp["code"], emp["email"],
            s1, s2, s3,
            s1_amt, s2_amt, s3_amt,
            grand
        ])

        for col in ws[row_idx]:
            col.alignment = center
            col.border = border

        row_idx += 1

    # 🔥 GRAND TOTAL ROW
    ws.append([
        "Grand Total", "", "",
        totals["S1"], totals["S2"], totals["S3"],
        totals["S1_amt"], totals["S2_amt"], totals["S3_amt"],
        totals["grand"]
    ])

    for cell in ws[row_idx]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # AUTO WIDTH
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=shift_allowance_{month}_{year}.xlsx"
        }
    )

@router.get("/dashboard")
def dashboard(month: int, year: int, db: Session = Depends(get_db)):

    roster = db.query(Roster).filter(
        Roster.month == month,
        Roster.year == year
    ).order_by(Roster.id.desc()).first()

    if not roster:
        return {}

    entries = db.query(RosterEntry).filter_by(roster_id=roster.id).all()

    from collections import defaultdict
    result = defaultdict(lambda: {
        "name": "",
        "S1": 0, "S2": 0, "S3": 0,
        "WO": 0, "LV": 0, "CO": 0, "GH": 0
    })

    daily = defaultdict(int)

    for e in entries:

        emp = db.query(Employee).filter_by(id=e.employee_id).first()

        if not emp:
            continue

        result[e.employee_id]["name"] = emp.name

        if e.shift_id:
            shift = db.query(Shift).filter_by(id=e.shift_id).first()
            code = shift.shift_code

            # 🔹 TOP + DAILY (ONLY S1 S2 S3)
            if code in ["S1", "S2", "S3"]:
                result[e.employee_id][code] += 1
                daily[str(e.date)] += 1

            # 🔹 LEAVE
            if code in ["WO", "LV", "CO", "GH"]:
                result[e.employee_id][code] += 1

    # 🔹 TOP EMPLOYEES
    top = []
    for emp in result.values():
        total = emp["S1"] + emp["S2"] + emp["S3"]
        top.append({"name": emp["name"], "total": total})

    # 🔹 SHIFT DISTRIBUTION
    shift_dist = {"S1":0,"S2":0,"S3":0}
    leave_dist = {"WO":0,"LV":0,"CO":0,"GH":0}

    for emp in result.values():
        for s in ["S1","S2","S3"]:
            shift_dist[s] += emp[s]

        for l in leave_dist:
            leave_dist[l] += emp[l]

    return {
        "top": sorted(top, key=lambda x: x["total"], reverse=True),
        "shift_distribution": shift_dist,
        "leave_distribution": leave_dist,
        "daily": daily
    }